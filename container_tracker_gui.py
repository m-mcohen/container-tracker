#!/usr/bin/env python3
"""
Container Tracker — legacy tkinter entry point.

Logic helpers live in the container_tracker.core package; this file is just the
v1 GUI shell (tkinter / CustomTkinter widgets, SetupDialog, ContainerTrackerApp).
The pywebview replacement is built alongside this file and will eventually
supersede it (see docs/MIGRATION_PROMPT.md).
"""

import logging
import os
import sys
import threading
import webbrowser
from datetime import datetime
from pathlib import Path

try:
    import customtkinter as ctk
    HAS_CTK = True
except ImportError:
    HAS_CTK = False
    import tkinter as tk
    from tkinter import ttk

from tkinter import messagebox, filedialog, END, StringVar
import tkinter.ttk as ttk_mod
import requests

try:
    from PIL import Image, ImageTk
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
HAS_OPENPYXL = True

from container_tracker.core import config as ct_config
from container_tracker.core import credentials as ct_credentials
from container_tracker.core.api import ShipsGoClient, resolve_scac
from container_tracker.core.config import (
    CONFIG_FILE, DATA_DIR, LOG_FILE, TRACKING_DB_FILE,
    is_first_run, load_config, load_json, save_config, save_json,
)
from container_tracker.core.constants import (
    ACCENT, API_BASE, APP_NAME, APP_SHORT_NAME, CARRIER_NAMES,
    CARRIER_SCAC_MAP, CONTAINER_COL_KEYWORDS, GITHUB_REPO, __version__,
)
from container_tracker.core.credentials import (
    KEYRING_SERVICE, KEYRING_USER, LEGACY_KEYRING_SERVICE,
    get_api_token, set_api_token,
)
from container_tracker.core.excel import (
    TRACKING_COL_MAP, create_template_excel, find_container_column,
    find_or_create_tracking_columns, read_containers_from_excel,
    update_excel_with_tracking,
)
from container_tracker.core.status import (
    API_KEY_PATTERN, EMAIL_PATTERN, extract_fields, validate_setup_fields,
)
from container_tracker.core.updates import check_for_update_async
from container_tracker.core.util import (
    EST, app_icon_path, now_est, now_est_short, open_in_explorer, resource_path,
)


logger = logging.getLogger(__name__)
LOGO_PATH = app_icon_path()  # window icon only; not rendered in-app

LIGHT = {
    "bg": "#ECEAE5", "card": "#FFFFFF", "input": "#F7F6F3",
    "primary": "#111111", "secondary": "#333333", "muted": "#666666",
    "hint": "#999999", "border": "#D0CEC8",
    "green": ACCENT, "green_dark": "#1D4ED8", "green_light": "#DBEAFE",
    "blue": "#1565C0", "blue_light": "#E3F2FD",
    "btn_bg": "#DDDBD5", "btn_text": "#333333",
    "thead": "#F2F0EC", "log_bg": "#FFFFFF",
    "sail_bg": "#E3F2FD", "sail_fg": "#0D47A1",
    "disc_bg": "#DBEAFE", "disc_fg": "#1E40AF",
    "stat_bg": "#FFFFFF",
}
DARK = {
    "bg": "#1A1A1F", "card": "#242429", "input": "#1E1E24",
    "primary": "#F0EDE8", "secondary": "#CCCCCC", "muted": "#909090",
    "hint": "#606060", "border": "#3A3A42",
    "green": ACCENT, "green_dark": "#1D4ED8", "green_light": "#1E2A44",
    "blue": "#64B5F6", "blue_light": "#1A2A3A",
    "btn_bg": "#333338", "btn_text": "#CCCCCC",
    "thead": "#1E1E24", "log_bg": "#1E1E24",
    "sail_bg": "#1A2A3A", "sail_fg": "#64B5F6",
    "disc_bg": "#1A3A24", "disc_fg": "#66BB6A",
    "stat_bg": "#242429",
}

class SetupDialog:
    """Modal first-run / settings editor for company_name + api_key + contact_email."""
    def __init__(self, parent, initial_company="", initial_api_key="", initial_email="",
                 mode="first_run", extra_info=None):
        """
        mode: 'first_run' (blocking, no cancel, × quits app) or 'settings' (has Save + Cancel).
        extra_info: optional dict with keys version, data_dir (Path), github_repo — shown read-only in settings.
        """
        self.mode = mode
        self.result = None  # dict or None on cancel
        self.parent = parent
        self._build(initial_company, initial_api_key, initial_email, extra_info or {})

    def _build(self, company, api_key, email, extra):
        title = "Welcome to Container Tracker" if self.mode == "first_run" else "Settings"
        if HAS_CTK:
            self.win = ctk.CTkToplevel(self.parent)
        else:
            self.win = tk.Toplevel(self.parent)
        self.win.title(title)
        self.win.resizable(False, False)
        self.win.transient(self.parent)
        try:
            self.win.iconbitmap(str(LOGO_PATH))
        except Exception:
            pass

        # × handler
        if self.mode == "first_run":
            self.win.protocol("WM_DELETE_WINDOW", self._exit_app)
        else:
            self.win.protocol("WM_DELETE_WINDOW", self._on_cancel)

        body = ctk.CTkFrame(self.win, fg_color="transparent") if HAS_CTK else tk.Frame(self.win)
        body.pack(padx=24, pady=20, fill="both", expand=True)

        if self.mode == "first_run":
            intro = "Let's get you set up. You can change these later in Settings."
        else:
            intro = "Update your company info and API credentials."
        (ctk.CTkLabel(body, text=intro, font=("Segoe UI", 12), justify="left") if HAS_CTK
         else tk.Label(body, text=intro, font=("Segoe UI", 10), justify="left")).pack(anchor="w", pady=(0, 14))

        self.company_var = StringVar(value=company)
        self.api_var = StringVar(value=api_key)
        self.email_var = StringVar(value=email)

        self._field(body, "Company name", self.company_var, width=360)
        self._field(body, "ShipsGo API key", self.api_var, width=360, mono=True,
                    hint="Find your key at shipsgo.com \u2192 Dashboard \u2192 Integrations \u2192 ShipsGo API")
        self._field(body, "Contact email", self.email_var, width=360)

        for v in (self.company_var, self.api_var, self.email_var):
            v.trace_add("write", lambda *_: self._update_save_state())

        if extra:
            sep = (ctk.CTkFrame(body, height=1, fg_color="#CCCCCC") if HAS_CTK
                   else tk.Frame(body, height=1, bg="#CCCCCC"))
            sep.pack(fill="x", pady=(12, 10))
            self._info_row(body, "App version", extra.get("version", ""))
            dd = extra.get("data_dir")
            if dd:
                self._info_row(body, "Data folder", str(dd), clickable=lambda: open_in_explorer(dd))
            self._info_row(body, "GitHub repo", extra.get("github_repo", ""))

        # Error label
        self.err_var = StringVar(value="")
        err = (ctk.CTkLabel(body, textvariable=self.err_var, text_color="#D32F2F",
                            font=("Segoe UI", 11), justify="left") if HAS_CTK
               else tk.Label(body, textvariable=self.err_var, fg="#D32F2F", font=("Segoe UI", 9), justify="left"))
        err.pack(anchor="w", pady=(6, 0))

        # Buttons
        btns = ctk.CTkFrame(body, fg_color="transparent") if HAS_CTK else tk.Frame(body)
        btns.pack(fill="x", pady=(14, 0))

        if self.mode == "settings":
            if HAS_CTK:
                self.cancel_btn = ctk.CTkButton(btns, text="Cancel", width=100, command=self._on_cancel,
                                                fg_color="#DDDBD5", text_color="#333333", hover_color="#C7C4BD")
            else:
                self.cancel_btn = tk.Button(btns, text="Cancel", width=12, command=self._on_cancel)
            self.cancel_btn.pack(side="right", padx=(8, 0))

        if HAS_CTK:
            self.save_btn = ctk.CTkButton(btns, text="Save", width=120, command=self._on_save,
                                          fg_color=ACCENT, hover_color="#1D4ED8", text_color="white",
                                          font=("Segoe UI", 12, "bold"))
        else:
            self.save_btn = tk.Button(btns, text="Save", width=14, command=self._on_save,
                                      bg=ACCENT, fg="white")
        self.save_btn.pack(side="right")

        self._update_save_state()
        self.win.update_idletasks()
        self._center()
        self.win.grab_set()
        self.win.focus_force()

    def _field(self, parent, label, var, width=340, mono=False, hint=None):
        (ctk.CTkLabel(parent, text=label, font=("Segoe UI", 12, "bold"), anchor="w") if HAS_CTK
         else tk.Label(parent, text=label, font=("Segoe UI", 10, "bold"), anchor="w")).pack(anchor="w", pady=(4, 2))
        font = ("Consolas", 12) if mono else ("Segoe UI", 12)
        if HAS_CTK:
            entry = ctk.CTkEntry(parent, textvariable=var, width=width, font=font, corner_radius=8)
        else:
            entry = tk.Entry(parent, textvariable=var, width=width // 8, font=font)
        entry.pack(anchor="w", pady=(0, 2))
        if hint:
            (ctk.CTkLabel(parent, text=hint, font=("Segoe UI", 10), text_color="#666666", anchor="w") if HAS_CTK
             else tk.Label(parent, text=hint, font=("Segoe UI", 9), fg="#666666", anchor="w")).pack(anchor="w", pady=(0, 8))
        else:
            (ctk.CTkFrame(parent, height=4, fg_color="transparent") if HAS_CTK
             else tk.Frame(parent, height=4)).pack(pady=(0, 4))

    def _info_row(self, parent, label, value, clickable=None):
        row = ctk.CTkFrame(parent, fg_color="transparent") if HAS_CTK else tk.Frame(parent)
        row.pack(anchor="w", fill="x", pady=1)
        (ctk.CTkLabel(row, text=label + ":", font=("Segoe UI", 10, "bold"), width=110, anchor="w") if HAS_CTK
         else tk.Label(row, text=label + ":", font=("Segoe UI", 9, "bold"), width=14, anchor="w")).pack(side="left")
        text_color = ACCENT if clickable else "#333333"
        cursor = "hand2" if clickable else ""
        if HAS_CTK:
            lbl = ctk.CTkLabel(row, text=value, font=("Segoe UI", 10), text_color=text_color,
                               cursor=cursor, anchor="w")
        else:
            lbl = tk.Label(row, text=value, font=("Segoe UI", 9), fg=text_color, cursor=cursor, anchor="w")
        lbl.pack(side="left")
        if clickable:
            lbl.bind("<Button-1>", lambda _e: clickable())

    def _center(self):
        self.win.update_idletasks()
        w = self.win.winfo_width(); h = self.win.winfo_height()
        sw = self.win.winfo_screenwidth(); sh = self.win.winfo_screenheight()
        x = (sw - w) // 2; y = (sh - h) // 3
        self.win.geometry(f"+{x}+{y}")

    def _update_save_state(self):
        valid = (self.company_var.get().strip() and self.api_var.get().strip() and self.email_var.get().strip()
                 and API_KEY_PATTERN.match(self.api_var.get().strip())
                 and EMAIL_PATTERN.match(self.email_var.get().strip()))
        state = "normal" if valid else "disabled"
        try:
            if HAS_CTK:
                self.save_btn.configure(state=state)
            else:
                self.save_btn.config(state=state)
        except Exception:
            pass

    def _on_save(self):
        err = validate_setup_fields(self.company_var.get(), self.api_var.get(), self.email_var.get())
        if err:
            self.err_var.set(err); return
        self.result = {
            "company_name": self.company_var.get().strip(),
            "api_key": self.api_var.get().strip(),
            "contact_email": self.email_var.get().strip(),
        }
        self.win.grab_release(); self.win.destroy()

    def _on_cancel(self):
        self.result = None
        self.win.grab_release(); self.win.destroy()

    def _exit_app(self):
        # First-run × quits app cleanly.
        try:
            self.win.grab_release(); self.win.destroy()
        except Exception:
            pass
        try:
            self.parent.destroy()
        except Exception:
            pass
        sys.exit(0)

_BaseRoot = ctk.CTk if HAS_CTK else tk.Tk


class ContainerTrackerApp(_BaseRoot):
    def __init__(self, cfg: dict | None = None):
        if HAS_CTK:
            ctk.set_appearance_mode("light")
        super().__init__()

        self.wm_attributes('-alpha', 0.0)

        self.update_idletasks()

        self.root = self  # alias so the rest of the class can keep using self.root.*

        # Boot/migrations are now run by the entry point (see __main__ below) or
        # by core.config.boot(). Caller passes a pre-loaded config; fall back to
        # load_config() for callers that haven't been updated yet.
        self.config = cfg if cfg is not None else load_config()
        self.is_dark=self.config.get("dark_mode",False)
        self.T=DARK if self.is_dark else LIGHT
        self.db=load_json(TRACKING_DB_FILE,{})
        self.client=None; self.themed_widgets=[]
        self.update_banner=None

        if HAS_CTK:
            self.configure(fg_color=self.T["bg"])
        else:
            self.configure(bg=self.T["bg"])
        self.title(APP_NAME)

        # Geometry is deferred to the reveal block below — calling self.geometry() on a
        # CTk root in 'withdrawn' state (CTk's default) is a silent no-op. minsize is fine
        # to set here; it just constrains future resize operations.
        self.minsize(860,650)

        if LOGO_PATH.exists() and LOGO_PATH.suffix.lower() == ".ico":
            def _set_icon():
                try: self.iconbitmap(str(LOGO_PATH))
                except Exception: pass
            self.after(200, _set_icon)
            self.after(600, _set_icon)
            self.after(1200, _set_icon)

        first_run = is_first_run(self.config)
        if first_run:
            dlg = SetupDialog(self, mode="first_run")
            self.wait_window(dlg.win)
            if not dlg.result:
                self.destroy()
                sys.exit(0)
            self.config["company_name"] = dlg.result["company_name"]
            self.config["contact_email"] = dlg.result["contact_email"]
            save_config(self.config)
            set_api_token(dlg.result["api_key"])
            self.api_key_cached = dlg.result["api_key"]
        else:
            self.api_key_cached = get_api_token()

        self._apply_window_title()
        self.build_ui(); self.load_table_data(); self.update_stats()

        # Reveal. Order matters: deiconify flips WM state 'withdrawn' → 'normal', so
        # geometry/alpha calls that follow actually take effect against a mapped window.
        self.deiconify()

        self.geometry("1020x800")

        self.wm_attributes('-alpha', 1.0)

        self.lift()
        self.update_idletasks()

        check_for_update_async(lambda tag, url: self.after(0, lambda: self.show_update_banner(tag, url)))

    def _apply_window_title(self):
        company = self.config.get("company_name", "").strip()
        self.root.title(f"{APP_NAME} \u2014 {company}" if company else APP_NAME)

    def show_update_banner(self, new_version: str, release_url: str):
        if self.update_banner is not None:
            return
        T=self.T
        if HAS_CTK:
            bar=ctk.CTkFrame(self.root, fg_color=T["sail_bg"], corner_radius=0, height=34)
            bar.pack(fill="x", side="top", before=self.root.winfo_children()[0])
            msg=ctk.CTkLabel(bar, text=f"Version {new_version} available \u2014 click to download",
                             text_color=T["sail_fg"], font=("Segoe UI",12,"bold"), cursor="hand2",
                             fg_color="transparent")
            msg.pack(side="left", padx=12, pady=4)
            close=ctk.CTkButton(bar, text="\u00d7", width=28, height=24, corner_radius=6,
                                fg_color="transparent", hover_color=T["border"],
                                text_color=T["sail_fg"], font=("Segoe UI",14,"bold"),
                                command=self._dismiss_update_banner)
            close.pack(side="right", padx=6, pady=4)
        else:
            bar=tk.Frame(self.root, bg=T["sail_bg"])
            bar.pack(fill="x", side="top", before=self.root.winfo_children()[0])
            msg=tk.Label(bar, text=f"Version {new_version} available \u2014 click to download",
                         bg=T["sail_bg"], fg=T["sail_fg"], font=("Segoe UI",10,"bold"), cursor="hand2")
            msg.pack(side="left", padx=12, pady=4)
            close=tk.Button(bar, text="\u00d7", bg=T["sail_bg"], fg=T["sail_fg"],
                            relief="flat", command=self._dismiss_update_banner)
            close.pack(side="right", padx=6, pady=2)
        msg.bind("<Button-1>", lambda _e: webbrowser.open(release_url))
        self.update_banner=bar

    def _dismiss_update_banner(self):
        if self.update_banner is not None:
            try: self.update_banner.destroy()
            except Exception: pass
            self.update_banner=None

    def _reg(self,w,role):
        self.themed_widgets.append((w,role)); return w

    def apply_theme(self):
        T=self.T
        for w,role in self.themed_widgets:
            try:
                if not HAS_CTK: continue
                m={"bg":{"fg_color":T["bg"]},"card":{"fg_color":T["card"],"border_color":T["border"]},
                   "stat_card":{"fg_color":T["stat_bg"],"border_color":T["border"]},
                   "input":{"fg_color":T["input"],"border_color":T["border"],"text_color":T["primary"]},
                   "label_primary":{"text_color":T["primary"]},"label_secondary":{"text_color":T["secondary"]},
                   "label_muted":{"text_color":T["muted"]},"label_hint":{"text_color":T["hint"]},
                   "label_green":{"text_color":T["green"]},
                   "btn":{"fg_color":T["btn_bg"],"text_color":T["btn_text"],"hover_color":T["border"]},
                   "btn_outline":{"fg_color":"transparent","border_color":T["green"],"text_color":T["green"],"hover_color":T["green_light"]},
                   "btn_green":{"fg_color":T["green"],"hover_color":T["green_dark"]},
                   "btn_red":{"fg_color":"#D32F2F","hover_color":"#B71C1C"},
                   "log":{"fg_color":T["log_bg"],"text_color":T["secondary"],"border_color":T["border"]},
                   "combo":{"fg_color":T["input"],"border_color":T["border"],"button_color":T["btn_bg"],
                            "dropdown_fg_color":T["card"],"text_color":T["primary"]},
                   "stat_value":{"text_color":T["primary"]},"stat_label":{"text_color":T["muted"]}}
                if role in m: w.configure(**m[role])
            except: pass
        if HAS_CTK: self.root.configure(fg_color=T["bg"])
        s=ttk_mod.Style()
        s.configure("Custom.Treeview",background=T["card"],fieldbackground=T["card"],
                    foreground=T["primary"],rowheight=34,font=("Segoe UI",11),borderwidth=0)
        s.configure("Custom.Treeview.Heading",background=T["thead"],foreground=T["muted"],
                    font=("Segoe UI",10),borderwidth=0,relief="flat")
        s.map("Custom.Treeview",background=[("selected",T["green_light"])],
              foreground=[("selected",T["primary"])])

    def toggle_theme(self):
        self.is_dark=not self.is_dark; self.T=DARK if self.is_dark else LIGHT
        self.config["dark_mode"]=self.is_dark; save_config(self.config)
        self.apply_theme()
        if HAS_CTK:
            if self.is_dark: self.theme_switch.select()
            else: self.theme_switch.deselect()

    def _f(self,p,role="bg"):
        w=ctk.CTkFrame(p,fg_color="transparent") if HAS_CTK else tk.Frame(p,bg=self.T["bg"])
        return self._reg(w,role)
    def _card(self,p):
        w=ctk.CTkFrame(p,fg_color=self.T["card"],corner_radius=12,border_width=1,border_color=self.T["border"]) if HAS_CTK else tk.Frame(p,bg=self.T["card"],bd=1,relief="solid")
        return self._reg(w,"card")
    def _l(self,p,text,role="label_primary",size=13,bold=False):
        wt="bold" if bold else "normal"
        cm={"label_primary":"primary","label_secondary":"secondary","label_muted":"muted","label_hint":"hint","label_green":"green"}
        c=self.T[cm.get(role,"primary")]
        w=ctk.CTkLabel(p,text=text,text_color=c,font=("Segoe UI",size,wt),fg_color="transparent") if HAS_CTK else tk.Label(p,text=text,fg=c,bg=self.T["bg"],font=("Segoe UI",size,wt))
        return self._reg(w,role)
    def _btn(self,p,text,cmd,role="btn"):
        if HAS_CTK:
            kw={"text":text,"command":cmd,"corner_radius":8}
            if role=="btn_green": kw.update(fg_color=self.T["green"],hover_color=self.T["green_dark"],text_color="white",font=("Segoe UI",13,"bold"),height=38)
            elif role=="btn_outline": kw.update(fg_color="transparent",hover_color=self.T["green_light"],text_color=self.T["green"],border_width=1,border_color=self.T["green"],font=("Segoe UI",12),height=34)
            elif role=="btn_red": kw.update(fg_color="#D32F2F",hover_color="#B71C1C",text_color="white",font=("Segoe UI",11),height=30,width=100)
            else: kw.update(fg_color=self.T["btn_bg"],hover_color=self.T["border"],text_color=self.T["btn_text"],font=("Segoe UI",12),height=34)
            w=ctk.CTkButton(p,**kw)
        else:
            bg=self.T["green"] if role=="btn_green" else ("#D32F2F" if role=="btn_red" else self.T["btn_bg"])
            fg="white" if role in ("btn_green","btn_red") else self.T["btn_text"]
            w=tk.Button(p,text=text,command=cmd,bg=bg,fg=fg,font=("Segoe UI",11),relief="flat",padx=12,pady=4)
        return self._reg(w,role)
    def _e(self,p,**kw):
        w=ctk.CTkEntry(p,fg_color=self.T["input"],border_color=self.T["border"],text_color=self.T["primary"],corner_radius=8,**kw) if HAS_CTK else tk.Entry(p,bg=self.T["input"],fg=self.T["primary"],relief="solid",bd=1)
        return self._reg(w,"input")

    def build_ui(self):
        T=self.T
        # Header
        hdr=self._f(self.root); hdr.pack(fill="x",padx=20,pady=(14,6))
        left=self._f(hdr); left.pack(side="left")
        tf=self._f(left); tf.pack(side="left")
        self._l(tf,APP_NAME,size=18,bold=True).pack(anchor="w")
        company=self.config.get("company_name","").strip()
        self.company_label=self._l(tf, company, role="label_muted", size=11)
        self.company_label.pack(anchor="w")
        rt=self._f(hdr); rt.pack(side="right")
        self.status_label=self._l(rt,"",role="label_green",size=11)
        self.status_label.pack(side="left",padx=(0,14))
        if HAS_CTK:
            self.theme_switch=ctk.CTkSwitch(rt,text="",width=44,command=self.toggle_theme,
                fg_color=T["border"],progress_color=T["green"],button_color="#FFF",button_hover_color="#EEE")
            if self.is_dark: self.theme_switch.select()
            self.theme_switch.pack(side="left")
            self.settings_btn=ctk.CTkButton(rt, text="\u2699", width=32, height=32, corner_radius=8,
                fg_color="transparent", hover_color=T["border"], text_color=T["secondary"],
                font=("Segoe UI", 18), command=self.open_settings)
            self.settings_btn.pack(side="left", padx=(10, 0))
        else:
            self.settings_btn=tk.Button(rt, text="\u2699", command=self.open_settings,
                font=("Segoe UI", 14), relief="flat", bg=T["bg"], fg=T["secondary"])
            self.settings_btn.pack(side="left", padx=(10, 0))

        # Excel card
        ec=self._card(self.root); ec.pack(fill="x",padx=20,pady=(4,4))
        ei=self._f(ec,role="card"); ei.pack(fill="x",padx=16,pady=12)
        if HAS_CTK: ei.configure(fg_color=T["card"])
        self._l(ei,"Linked spreadsheet",role="label_muted",size=11).pack(anchor="w",pady=(0,4))
        er=self._f(ei,role="card"); er.pack(fill="x")
        if HAS_CTK: er.configure(fg_color=T["card"])
        self.excel_display=self._l(er,self.config.get("excel_path","") or "No file linked",role="label_green",size=11)
        self.excel_display.pack(side="left",padx=(0,12))
        self._btn(er,"Browse...",self.browse_excel).pack(side="left",padx=3)
        self._btn(er,"Create Template",self.create_template).pack(side="left",padx=3)
        self._btn(er,"Open in Excel",self.open_excel).pack(side="left",padx=3)

        # Summary cards
        sf=self._f(self.root); sf.pack(fill="x",padx=20,pady=(6,4))
        self.stat_frames={}
        for key,label in [("total","Tracked"),("sailing","Sailing"),("arrived","Arrived"),("delayed","Delayed")]:
            card=ctk.CTkFrame(sf,fg_color=T["stat_bg"],corner_radius=10,border_width=1,border_color=T["border"],height=60) if HAS_CTK else tk.Frame(sf,bg=T["stat_bg"],bd=1,relief="solid")
            self._reg(card,"stat_card")
            card.pack(side="left",fill="x",expand=True,padx=(0 if key=="total" else 4,0))
            sl=self._l(card,label,role="stat_label",size=10); sl.pack(anchor="w",padx=12,pady=(8,0))
            color=T["primary"]
            if key=="sailing": color=T["blue"]
            elif key=="arrived": color=T["green"]
            elif key=="delayed": color="#D32F2F"
            sv=ctk.CTkLabel(card,text="0",text_color=color,font=("Segoe UI",22,"bold"),fg_color="transparent") if HAS_CTK else tk.Label(card,text="0",fg=color,bg=T["stat_bg"],font=("Segoe UI",22,"bold"))
            sv.pack(anchor="w",padx=12,pady=(0,8))
            self.stat_frames[key]=sv

        # Actions
        af=self._f(self.root); af.pack(fill="x",padx=20,pady=(6,4))
        self.refresh_btn=self._btn(af,"  Refresh All ETAs & Update Excel  ",self.refresh_data,role="btn_green")
        self.refresh_btn.pack(side="left",padx=(0,8))
        self._btn(af,"Remove Selected",self.remove_container,role="btn_red").pack(side="left",padx=(0,16))
        self._l(af,"Add:",role="label_muted",size=11).pack(side="left",padx=(12,4))
        self.container_var=StringVar()
        self._e(af,textvariable=self.container_var,width=130).pack(side="left",padx=(0,4))
        self.carrier_var=StringVar(value="MAERSK LINE")
        if HAS_CTK:
            cc=ctk.CTkComboBox(af,values=CARRIER_NAMES,variable=self.carrier_var,width=130,
                fg_color=T["input"],border_color=T["border"],button_color=T["btn_bg"],
                dropdown_fg_color=T["card"],text_color=T["primary"],corner_radius=8)
            self._reg(cc,"combo")
        else: cc=ttk.Combobox(af,textvariable=self.carrier_var,values=CARRIER_NAMES,width=14,state="readonly")
        cc.pack(side="left",padx=(0,4))
        self._btn(af,"Add & Track",self.add_container,role="btn_outline").pack(side="left")

        # Table
        tbf=self._f(self.root); tbf.pack(fill="both",expand=True,padx=20,pady=(4,4))
        s=ttk_mod.Style(); s.theme_use("clam")
        s.configure("Custom.Treeview",background=T["card"],fieldbackground=T["card"],
                    foreground=T["primary"],rowheight=34,font=("Segoe UI",11),borderwidth=0)
        s.configure("Custom.Treeview.Heading",background=T["thead"],foreground=T["muted"],
                    font=("Segoe UI",10),borderwidth=0,relief="flat")
        s.map("Custom.Treeview",background=[("selected",T["green_light"])],foreground=[("selected",T["primary"])])
        cols=("container","carrier","status","orig_eta","eta","delay","route","vessel","transit")
        self.tree=ttk_mod.Treeview(tbf,columns=cols,show="headings",height=8,style="Custom.Treeview")
        for cid,hd,w in [("container","Container #",115),("carrier","Carrier",90),("status","Status",90),
            ("orig_eta","Original ETA",90),("eta","Current ETA",90),("delay","Delay",80),
            ("route","Route",180),("vessel","Vessel",120),("transit","Transit",60)]:
            self.tree.heading(cid,text=hd); self.tree.column(cid,width=w,minwidth=45)
        if HAS_CTK:
            sb=ctk.CTkScrollbar(tbf,command=self.tree.yview,fg_color="transparent",
                button_color=T["border"],button_hover_color=T["muted"])
        else:
            sb=ttk_mod.Scrollbar(tbf,orient="vertical",command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side="left",fill="both",expand=True); sb.pack(side="right",fill="y")

        # Log
        lf=self._f(self.root); lf.pack(fill="x",padx=20,pady=(2,2))
        self._l(lf,"Activity log",role="label_hint",size=10).pack(anchor="w",pady=(0,2))
        if HAS_CTK:
            self.log_text=ctk.CTkTextbox(self.root,height=80,fg_color=T["log_bg"],text_color=T["secondary"],
                font=("Consolas",11),corner_radius=8,border_width=1,border_color=T["border"])
            self._reg(self.log_text,"log")
        else:
            self.log_text=tk.Text(self.root,height=4,font=("Consolas",9),bg=T["log_bg"],fg=T["secondary"],relief="solid",bd=1)
        self.log_text.pack(fill="x",padx=20,pady=(0,8))

        # Footer
        ff=self._f(self.root); ff.pack(fill="x",padx=20,pady=(0,8))
        self._l(ff,"Powered by ShipsGo API",role="label_hint",size=9).pack(side="left")
        self._l(ff,"Refreshes are free & unlimited \u2022 All times EST",role="label_hint",size=9).pack(side="right")

    def log(self,msg):
        self.log_text.insert(END,f"[{now_est_short()}] {msg}\n"); self.log_text.see(END); logger.info(msg)
    def set_status(self,msg):
        if HAS_CTK: self.status_label.configure(text=msg)
        else: self.status_label.config(text=msg)
        self.root.update_idletasks()
    def _dis(self):
        if HAS_CTK: self.refresh_btn.configure(state="disabled")
        else: self.refresh_btn.config(state="disabled")
    def _en(self):
        if HAS_CTK: self.refresh_btn.configure(state="normal")
        else: self.refresh_btn.config(state="normal")

    def update_stats(self):
        total=len(self.db); sailing=0; arrived=0; delayed=0
        for _,r in self.db.items():
            st=str(r.get("status","")).upper()
            dd=str(r.get("delay_days",""))
            if st=="SAILING": sailing+=1
            elif st in ("ARRIVED","DISCHARGED","DELIVERED","GATE_OUT"): arrived+=1
            if dd.startswith("+") and st=="SAILING": delayed+=1
        for k,v in [("total",total),("sailing",sailing),("arrived",arrived),("delayed",delayed)]:
            if HAS_CTK: self.stat_frames[k].configure(text=str(v))
            else: self.stat_frames[k].config(text=str(v))

    def load_table_data(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        for key,rec in sorted(self.db.items()):
            tp=rec.get("transit_pct","")
            if tp!="": tp=f"{tp}%"
            pol=rec.get("pol",""); pod=rec.get("pod","")
            route=f"{pol} \u2192 {pod}" if pol and pod else pol or pod or ""
            self.tree.insert("",END,iid=key,values=(
                rec.get("container_number") or key, rec.get("carrier",rec.get("shipping_line","")),
                rec.get("status",""), rec.get("original_eta",""), rec.get("eta",""),
                rec.get("delay_days",""), route, rec.get("vessel",""), tp))
        self.update_stats()

    def open_settings(self):
        extra = {"version": __version__, "data_dir": DATA_DIR, "github_repo": GITHUB_REPO}
        dlg = SetupDialog(self.root,
                          initial_company=self.config.get("company_name",""),
                          initial_api_key=self.api_key_cached or get_api_token(),
                          initial_email=self.config.get("contact_email",""),
                          mode="settings",
                          extra_info=extra)
        self.root.wait_window(dlg.win)
        if not dlg.result:
            return
        self.config["company_name"] = dlg.result["company_name"]
        self.config["contact_email"] = dlg.result["contact_email"]
        save_config(self.config)
        if dlg.result["api_key"] != self.api_key_cached:
            set_api_token(dlg.result["api_key"])
            self.api_key_cached = dlg.result["api_key"]
            self.client = None  # force rebuild with new token
        self._apply_window_title()
        company = self.config.get("company_name", "").strip()
        try:
            if HAS_CTK: self.company_label.configure(text=company)
            else: self.company_label.config(text=company)
        except Exception: pass
        self.log("Settings updated.")

    def get_client(self):
        key=(self.api_key_cached or get_api_token()).strip()
        if not key:
            messagebox.showwarning("Missing API Key","Open Settings (\u2699) and enter your ShipsGo API key.\n\n1. Go to shipsgo.com\n2. Dashboard > Integrations > ShipsGo API\n3. Copy your token"); return None
        if self.client is None: self.client=ShipsGoClient(key)
        return self.client

    def browse_excel(self):
        p=filedialog.askopenfilename(title="Select spreadsheet",filetypes=[("Excel","*.xlsx"),("All","*.*")])
        if p:
            self.config["excel_path"]=p; save_config(self.config)
            if HAS_CTK: self.excel_display.configure(text=p)
            else: self.excel_display.config(text=p)
            self.log(f"Linked: {Path(p).name}")
    def create_template(self):
        p=filedialog.asksaveasfilename(title="Save template",defaultextension=".xlsx",initialfile="Container_Tracking.xlsx",filetypes=[("Excel","*.xlsx")])
        if p:
            try:
                create_template_excel(p); self.config["excel_path"]=p; save_config(self.config)
                if HAS_CTK: self.excel_display.configure(text=p)
                else: self.excel_display.config(text=p)
                self.log(f"Template created: {Path(p).name}")
                messagebox.showinfo("Template Created","Template saved as an Excel Table.\n\nReplace samples with real containers, then Refresh.\nNew rows auto-inherit table formatting.")
                os.startfile(p)
            except Exception as e: messagebox.showerror("Error",str(e))
    def open_excel(self):
        p=self.config.get("excel_path","")
        if p and Path(p).exists(): os.startfile(p)
        else: messagebox.showinfo("No File","No Excel file linked.\n\nClick 'Browse...' or 'Create Template'.")

    def remove_container(self):
        sel=self.tree.selection()
        if not sel: messagebox.showinfo("No Selection","Select a container in the table first."); return
        cn=sel[0]
        rec=self.db.get(cn,{})
        status=str(rec.get("status","")).upper()
        is_done=status in ("DISCHARGED","DELIVERED","GATE_OUT","ARRIVED")

        if is_done:
            if not messagebox.askyesno("Remove Completed Shipment",
                f"Remove {cn}?\n\nStatus: {status}\n\n"
                "This shipment is complete. It will be permanently\n"
                "dismissed and won't reappear on future refreshes.\n\n"
                "It will remain in your Excel file."): return
            # Add to permanent dismissed list
            if "dismissed" not in self.config: self.config["dismissed"]=[]
            if cn not in self.config["dismissed"]:
                self.config["dismissed"].append(cn)
            save_config(self.config)
            if cn in self.db: del self.db[cn]
            save_json(TRACKING_DB_FILE,self.db); self.load_table_data()
            self.log(f"Dismissed {cn} (completed shipment, won't reappear)")
        else:
            if not messagebox.askyesno("Remove Active Shipment",
                f"Remove {cn} from the app?\n\nStatus: {status}\n\n"
                "This shipment is still active on ShipsGo.\n"
                "It WILL reappear on the next refresh.\n\n"
                "To permanently stop tracking, wait until\n"
                "the shipment is discharged, then remove it."): return
            if cn in self.db: del self.db[cn]
            save_json(TRACKING_DB_FILE,self.db); self.load_table_data()
            self.log(f"Removed {cn} from display (will reappear on next refresh)")

    def add_container(self):
        client=self.get_client()
        if not client: return
        cn=self.container_var.get().strip().upper(); cl=self.carrier_var.get().strip()
        if not cn: messagebox.showwarning("Missing","Enter a container number."); return
        if len(cn)!=11:
            if not messagebox.askyesno("Check Container #",f"Usually 11 chars (4 letters + 7 digits).\nYours: {cn} ({len(cn)} chars)\n\nContinue?"): return
        scac=resolve_scac(cl)
        if not messagebox.askyesno("Confirm Registration",
            f"Register {cn} with ShipsGo?\n\n"
            f"This will use 1 tracking credit (~$2 USD).\n"
            f"Credits are one-time per shipment \u2014 all future\n"
            f"refreshes are free and unlimited.\n\n"
            f"If the container is already tracked, no credit\n"
            f"will be charged."): return
        def _go():
            # Un-dismiss if previously removed
            dismissed=self.config.get("dismissed",[])
            if cn in dismissed:
                dismissed.remove(cn); self.config["dismissed"]=dismissed; save_config(self.config)
                self.log(f"Re-activated {cn} (was previously dismissed)")
            self.set_status("Registering..."); self.log(f"Adding {cn} ({cl})...")
            try:
                r=client.create_shipment(container_number=cn,carrier_scac=scac)
                if r.get("error")=="NOT_ENOUGH_CREDITS":
                    self.log(f"Not enough credits for {cn}")
                    self.root.after(0,lambda:messagebox.showerror("No Credits",
                        "Not enough ShipsGo credits.\n\n"
                        "To purchase more credits:\n"
                        "1. Go to shipsgo.com\n"
                        "2. Log into your dashboard\n"
                        "3. Click 'Buy Now' (starts at $20 for 10 credits)\n\n"
                        "Then come back and try again."))
                elif r.get("already_exists"):
                    self.log(f"{cn} already tracked on ShipsGo")
                    # Still add to local DB so it shows in app
                    if cn not in self.db:
                        self.db[cn]={"container_number":cn,"carrier":cl,"last_refreshed":None}
                        save_json(TRACKING_DB_FILE,self.db)
                else:
                    self.log(f"{cn} registered (1 credit used)")
                    # Add to local DB immediately so it shows in app
                    self.db[cn]={"container_number":cn,"carrier":cl,
                                 "shipment_id":r.get("id",""),"last_refreshed":None}
                    save_json(TRACKING_DB_FILE,self.db)
                    self.root.after(0,lambda:messagebox.showinfo("Added",
                        f"{cn} registered successfully.\n\n"
                        f"Full tracking data may take a few hours to appear.\n"
                        f"The container will be added to your Excel file on\n"
                        f"the next refresh."))
                self._do_refresh()
            except requests.ConnectionError:
                self.log("Connection error"); self.root.after(0,lambda:messagebox.showerror("No Connection","Check your internet connection."))
            except Exception as e:
                self.log(f"Error: {e}"); self.root.after(0,lambda:messagebox.showerror("Error",str(e)))
            finally: self.set_status(""); self.root.after(0,self._en)
        self._dis(); threading.Thread(target=_go,daemon=True).start()

    def refresh_data(self):
        if not self.get_client(): return
        def _t(): self._do_refresh(); self.set_status(""); self.root.after(0,self._en)
        self._dis(); threading.Thread(target=_t,daemon=True).start()

    def _do_refresh(self):
        client=self.get_client()
        if not client: return
        self.set_status("Fetching..."); self.log("Refreshing...")
        try:
            ships=client.list_shipments(); ac=len(ships)
            self.log(f"Found {ac} shipments on ShipsGo")
            if ac==0:
                self.log("WARNING: No shipments on your account")
                self.root.after(0,lambda:messagebox.showwarning("No Shipments","No shipments on ShipsGo.\n\nUse 'Add & Track' to register containers (1 credit each)."))
                return
            smap={}
            for s in ships:
                if not isinstance(s,dict): continue
                sid=s.get("id")
                if sid: smap[str(sid)]=s
                cn=(s.get("container_number") or "").upper()
                if cn: smap[cn]=s
            ep=self.config.get("excel_path","")
            if ep and Path(ep).exists():
                try:
                    ec=read_containers_from_excel(ep); self.log(f"Read {len(ec)} containers from Excel")
                    if len(ec)==0: self.log("WARNING: No containers found in spreadsheet")
                    dismissed=self.config.get("dismissed",[])
                    for c in ec:
                        if c not in self.db and c not in dismissed:
                            self.db[c]={"container_number":c,"last_refreshed":None}
                except PermissionError:
                    self.log("ERROR: Excel open - close it first")
                    self.root.after(0,lambda:messagebox.showerror("File In Use","Close Excel first, then Refresh.")); return
                except Exception as e: self.log(f"Excel read error: {e}")
            if not self.db and smap:
                dismissed=self.config.get("dismissed",[])
                for s in ships:
                    if not isinstance(s,dict): continue
                    cn=(s.get("container_number") or "").upper()
                    if not cn or cn in dismissed: continue
                    cr=s.get("carrier") or {}
                    self.db[cn]={"container_number":cn,"shipping_line":cr.get("name","") if isinstance(cr,dict) else "","shipment_id":s.get("id",""),"last_refreshed":None}
            matched=0; unmatched=0; delayed_sailing=0; unmatched_list=[]
            for key,rec in self.db.items():
                sid=str(rec.get("shipment_id","")); cn=rec.get("container_number","").upper()
                sh=smap.get(sid) or smap.get(cn)
                if sh:
                    fid=sh.get("id")
                    if fid:
                        try: self.set_status(f"Fetching {cn}..."); sh=client.get_shipment(fid); rec["shipment_id"]=fid
                        except: pass
                    fe=extract_fields(sh); rec.update(fe); rec["last_refreshed"]=now_est()
                    dd=fe.get("delay_days",""); st=fe.get("status","").upper()
                    delay_info=f" | DELAYED {dd}" if dd.startswith("+") else ""
                    if dd.startswith("+") and st=="SAILING": delayed_sailing+=1
                    self.log(f"  {cn}: {fe['status']} | ETA: {fe['eta']} | {fe['pol']} -> {fe['pod']}{delay_info}")
                    matched+=1
                else:
                    rec["last_refreshed"]=now_est(); self.log(f"  {cn}: not on ShipsGo yet")
                    unmatched+=1; unmatched_list.append(cn)
            save_json(TRACKING_DB_FILE,self.db); self.root.after(0,self.load_table_data)
            eu=0
            if ep and Path(ep).exists():
                try:
                    self.set_status("Updating Excel..."); eu=update_excel_with_tracking(ep,self.db)
                    self.log(f"Updated {eu} rows in Excel")
                except PermissionError:
                    self.log("Excel open - close it first")
                    self.root.after(0,lambda:messagebox.showwarning("File In Use","Close Excel, then Refresh."))
                except Exception as e: self.log(f"Excel error: {e}")
            self.log(f"--- DONE: {matched} matched, {unmatched} unmatched, {delayed_sailing} actively delayed, {eu} Excel rows updated ---")
            self.set_status(f"Refreshed {matched} containers \u2014 {now_est_short()} EST")

            # Show unmatched containers popup with option to register
            if unmatched_list:
                ul=list(unmatched_list)  # capture for lambda
                self.root.after(0, lambda: self._prompt_register_unmatched(ul))
            elif delayed_sailing>0:
                self.root.after(0,lambda:messagebox.showinfo("Delays Detected",f"{delayed_sailing} container(s) currently sailing are delayed.\n\nCheck the Delay column for details."))
        except requests.ConnectionError:
            self.log("Connection error"); self.root.after(0,lambda:messagebox.showerror("No Connection","Check your internet."))
        except requests.HTTPError as e:
            if "401" in str(e):
                self.log("Auth failed"); self.root.after(0,lambda:messagebox.showerror("Invalid API Key","API key rejected.\n\nRe-copy from shipsgo.com > Dashboard > Integrations."))
            else: self.log(f"API error: {e}"); self.root.after(0,lambda:messagebox.showerror("Error",str(e)))
        except Exception as e:
            self.log(f"Failed: {e}"); self.root.after(0,lambda:messagebox.showerror("Error",str(e)))

    def _prompt_register_unmatched(self, containers):
        """Show popup listing unmatched containers with option to register them."""
        container_list = "\n".join(f"  \u2022 {c}" for c in containers[:15])
        if len(containers) > 15:
            container_list += f"\n  ... and {len(containers)-15} more"

        result = messagebox.askyesno(
            f"{len(containers)} New Container(s) Found",
            f"The following containers are in your spreadsheet but not yet "
            f"tracked on ShipsGo:\n\n{container_list}\n\n"
            f"Would you like to register them now?\n\n"
            f"Cost: 1 credit per container (~$2 USD each)\n"
            f"Credits are one-time per shipment \u2014 all future\n"
            f"refreshes are free and unlimited.\n\n"
            f"Total: {len(containers)} credit(s) will be used.")

        if result:
            self._dis()
            threading.Thread(target=self._register_unmatched, args=(containers,), daemon=True).start()

    def _register_unmatched(self, containers):
        """Register unmatched containers with ShipsGo in background."""
        client = self.get_client()
        if not client:
            self._en(); return

        registered = 0; failed = 0; out_of_credits = False
        self.log(f"Registering {len(containers)} new containers...")

        for cn in containers:
            self.set_status(f"Registering {cn}...")
            try:
                r = client.create_shipment(container_number=cn)
                if r.get("error") == "NOT_ENOUGH_CREDITS":
                    out_of_credits = True
                    self.log(f"  {cn}: out of credits")
                    remaining = len(containers) - registered - failed
                    self.root.after(0, lambda rem=remaining: messagebox.showwarning(
                        "Out of Credits",
                        f"You ran out of ShipsGo credits.\n\n"
                        f"Registered: {registered}\n"
                        f"Remaining: {rem}\n\n"
                        f"To purchase more credits:\n"
                        f"1. Go to shipsgo.com\n"
                        f"2. Log into your dashboard\n"
                        f"3. Click 'Buy Now' (starts at $20 for 10 credits)\n\n"
                        f"Then come back and click Refresh to register\n"
                        f"the remaining containers."))
                    break
                elif r.get("already_exists"):
                    self.log(f"  {cn}: already on ShipsGo")
                    registered += 1
                else:
                    self.log(f"  {cn}: registered (1 credit)")
                    registered += 1
            except Exception as e:
                self.log(f"  {cn}: error - {e}")
                failed += 1

        if not out_of_credits:
            self.log(f"Registration complete: {registered} registered, {failed} failed")
            if registered > 0:
                self.root.after(0, lambda: messagebox.showinfo(
                    "Registration Complete",
                    f"{registered} container(s) registered successfully.\n\n"
                    f"Full tracking data may take a few hours to appear.\n"
                    f"Click Refresh periodically to check."))

        # Refresh to pick up newly registered containers
        self.log("Refreshing after registration...")
        self._do_refresh()
        self.root.after(0, self._en)

    def run(self): self.root.mainloop()

if __name__ == "__main__":
    ct_config.boot()
    ContainerTrackerApp(ct_config.load_config()).run()
