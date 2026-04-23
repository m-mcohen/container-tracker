#!/usr/bin/env python3
"""
Container ETA Tracker using ShipsGo API v2
==========================================
Tracks shipping containers across 160+ carriers via ShipsGo's API.
Outputs results to an Excel spreadsheet with ETA, status, and voyage details.

Setup:
  1. Sign up at https://shipsgo.com (free trial = 3 credits)
  2. Get your API token from Dashboard > Integrations > ShipsGo API
  3. Set in .env file: SHIPSGO_API_KEY=your_token_here
  4. Install dependencies: pip install requests openpyxl python-dotenv
"""

import json
import os
import sys
import time
import argparse
import logging
from datetime import datetime, timezone
from pathlib import Path

import requests

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
API_BASE = "https://api.shipsgo.com/v2"
CONTAINERS_FILE = "containers.json"
TRACKING_DB_FILE = "tracking_data.json"
OUTPUT_XLSX = "container_tracking_report.xlsx"
LOG_FILE = "tracker.log"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler(sys.stdout),
    ],
)
logger = logging.getLogger(__name__)


def get_api_key() -> str:
    key = os.environ.get("SHIPSGO_API_KEY", "")
    if not key:
        logger.error(
            "SHIPSGO_API_KEY not set. "
            "Export it or add to a .env file.\n"
            "  Sign up free at https://shipsgo.com to get your token."
        )
        sys.exit(1)
    return key


# ---------------------------------------------------------------------------
# ShipsGo API v2 Client
# ---------------------------------------------------------------------------
class ShipsGoClient:
    """Wrapper around ShipsGo REST API v2"""

    def __init__(self, token: str):
        self.token = token
        self.session = requests.Session()
        self.session.headers.update({
            "Accept": "application/json",
            "Content-Type": "application/json",
            "X-Shipsgo-User-Token": token,
        })

    def _log_response_error(self, resp, context=""):
        try:
            body = resp.text[:500]
        except Exception:
            body = "(unable to read response)"
        logger.error(f"  {context} HTTP {resp.status_code}: {body}")

    def create_shipment(self, container_number="", bl_number="",
                        carrier_scac="", reference=""):
        """Create an ocean shipment. Consumes 1 credit."""
        url = f"{API_BASE}/ocean/shipments"
        payload = {}
        if container_number:
            payload["container_number"] = container_number.strip().upper()
        if bl_number:
            payload["bl_number"] = bl_number.strip()
        if carrier_scac:
            payload["carrier_scac"] = carrier_scac.strip().upper()
        if reference:
            payload["reference"] = reference.strip()

        identifier = container_number or bl_number
        logger.info(f"POST create shipment: {identifier} (carrier: {carrier_scac or 'auto'})")
        resp = self.session.post(url, json=payload, timeout=30)

        if resp.status_code == 409:
            logger.info(f"  -> Already tracked (409 Conflict)")
            return {"already_exists": True, "identifier": identifier}
        if not resp.ok:
            self._log_response_error(resp, "Create shipment failed:")
            resp.raise_for_status()
        data = resp.json()
        logger.info(f"  -> Created shipment ID: {data.get('id', 'unknown')}")
        return data

    def list_shipments(self, take=100):
        """List all ocean shipments."""
        url = f"{API_BASE}/ocean/shipments"
        resp = self.session.get(url, params={"take": take}, timeout=30)
        if not resp.ok:
            self._log_response_error(resp, "List shipments failed:")
            resp.raise_for_status()
        return resp.json()

    def get_shipment(self, shipment_id):
        """Get full details for a single shipment."""
        url = f"{API_BASE}/ocean/shipments/{shipment_id}"
        resp = self.session.get(url, timeout=30)
        if not resp.ok:
            self._log_response_error(resp, f"Get shipment {shipment_id} failed:")
            resp.raise_for_status()
        return resp.json()

    def get_carriers(self):
        """List supported ocean carriers."""
        url = f"{API_BASE}/ocean/carriers"
        resp = self.session.get(url, timeout=30)
        if not resp.ok:
            self._log_response_error(resp, "Get carriers failed:")
            resp.raise_for_status()
        data = resp.json()
        return data.get("data", data) if isinstance(data, dict) else data


# ---------------------------------------------------------------------------
# SCAC Code Mapping
# ---------------------------------------------------------------------------
CARRIER_SCAC_MAP = {
    "MAERSK": "MAEU", "MAERSK LINE": "MAEU",
    "MSC": "MSCU", "CMA CGM": "CMDU",
    "HAPAG LLOYD": "HLCU", "HAPAG-LLOYD": "HLCU",
    "COSCO": "COSU", "EVERGREEN": "EGLV",
    "ONE": "ONEY", "YANG MING": "YMLU",
    "ZIM": "ZIMU", "HMM": "HDMU",
    "OOCL": "OOLU", "PIL": "PILU",
}

def resolve_scac(shipping_line):
    upper = shipping_line.strip().upper()
    if upper in CARRIER_SCAC_MAP:
        return CARRIER_SCAC_MAP[upper]
    if len(upper) == 4 and upper.isalpha():
        return upper
    return upper


# ---------------------------------------------------------------------------
# Local Data Store
# ---------------------------------------------------------------------------
def load_json(filepath, default=None):
    p = Path(filepath)
    if p.exists():
        with open(p) as f:
            return json.load(f)
    return default if default is not None else {}

def save_json(filepath, data):
    with open(filepath, "w") as f:
        json.dump(data, f, indent=2, default=str)

def load_containers():
    data = load_json(CONTAINERS_FILE, [])
    if not data:
        logger.warning(f"{CONTAINERS_FILE} not found or empty.")
    return data

def load_tracking_db():
    return load_json(TRACKING_DB_FILE, {})

def save_tracking_db(db):
    save_json(TRACKING_DB_FILE, db)


# ---------------------------------------------------------------------------
# Core Logic
# ---------------------------------------------------------------------------
def register_containers(client, containers, db):
    for entry in containers:
        container_num = entry.get("container_number", "").upper().strip()
        bl_num = entry.get("bl_number", "").strip()
        shipping_line = entry.get("shipping_line", "OTHERS")
        reference = entry.get("reference", "")
        key = container_num or bl_num
        if not key:
            continue
        if key in db and db[key].get("shipment_id"):
            logger.info(f"Already tracked: {key}")
            continue

        scac = resolve_scac(shipping_line)
        try:
            result = client.create_shipment(
                container_number=container_num,
                bl_number=bl_num,
                carrier_scac=scac,
                reference=reference,
            )
            db[key] = {
                "container_number": container_num,
                "bl_number": bl_num,
                "shipping_line": shipping_line,
                "carrier_scac": scac,
                "reference": reference,
                "shipment_id": result.get("id", ""),
                "registered_at": datetime.now(timezone.utc).isoformat(),
                "last_refreshed": None,
            }
            save_tracking_db(db)
            time.sleep(0.5)
        except requests.HTTPError as e:
            logger.error(f"Failed to register {key}: {e}")
        except Exception as e:
            logger.error(f"Unexpected error registering {key}: {e}")
    return db


def extract_fields(shipment):
    """Extract key fields from a v2 shipment response."""
    # Handle wrapper: API returns {"message": ..., "shipment": {...}}
    if "shipment" in shipment and isinstance(shipment["shipment"], dict):
        shipment = shipment["shipment"]

    fields = {"status": shipment.get("status", "Unknown"),
              "vessel": "", "pol": "", "pod": "", "eta": "",
              "etd": "", "carrier": "", "transit_pct": ""}

    # Carrier
    carrier = shipment.get("carrier") or {}
    if isinstance(carrier, dict):
        fields["carrier"] = carrier.get("name", carrier.get("scac", ""))

    # Route — v2 uses port_of_loading / port_of_discharge
    route = shipment.get("route") or {}

    pol = route.get("port_of_loading") or route.get("origin") or {}
    pol_loc = pol.get("location") or {}
    fields["pol"] = pol_loc.get("name", "")
    fields["etd"] = pol.get("date_of_loading", pol.get("date_of_dep", ""))

    pod = route.get("port_of_discharge") or route.get("destination") or {}
    pod_loc = pod.get("location") or {}
    fields["pod"] = pod_loc.get("name", "")
    fields["eta"] = pod.get("date_of_discharge", pod.get("date_of_eta", ""))

    fields["transit_pct"] = route.get("transit_percentage", "")

    # Vessel — get from last movement with a vessel
    containers = shipment.get("containers") or []
    if containers and isinstance(containers[0], dict):
        movements = containers[0].get("movements") or []
        # Find the most recent actual movement with a vessel
        for m in reversed(movements):
            if isinstance(m, dict) and m.get("vessel"):
                v = m["vessel"]
                if isinstance(v, dict) and v.get("name"):
                    fields["vessel"] = v["name"]
                    break

    # Clean up date strings — just show the date part
    for date_field in ("eta", "etd"):
        val = fields[date_field]
        if val and "T" in str(val):
            fields[date_field] = str(val).split("T")[0]

    return fields


def refresh_all(client, db):
    try:
        all_shipments = client.list_shipments()
        if isinstance(all_shipments, dict):
            logger.info(f"API response keys: {list(all_shipments.keys())}")
            all_shipments = all_shipments.get("shipments", all_shipments.get("data", []))
        logger.info(f"Fetched {len(all_shipments)} shipments from API")
        if all_shipments:
            first = all_shipments[0]
            logger.info(f"  First item type: {type(first).__name__}")
            if isinstance(first, dict):
                logger.info(f"  First item keys: {list(first.keys())[:10]}")
    except Exception as e:
        logger.error(f"Failed to list shipments: {e}")
        all_shipments = []

    shipment_map = {}
    for s in all_shipments:
        if not isinstance(s, dict):
            continue
        sid = s.get("id")
        if sid:
            shipment_map[str(sid)] = s
        # v2 API uses container_number at top level
        cnum = (s.get("container_number") or "").upper()
        if cnum:
            shipment_map[cnum] = s
        # Also check nested containers array if present
        for c in (s.get("containers") or []):
            if isinstance(c, dict):
                cn = c.get("number", "").upper()
                if cn:
                    shipment_map[cn] = s

    logger.info(f"  Shipment map keys: {list(shipment_map.keys())}")
    logger.info(f"  DB keys: {list(db.keys())}")

    # If DB is empty but API has shipments, populate DB from API
    if not db and shipment_map:
        logger.info("  DB is empty — populating from API shipments")
        for s in all_shipments:
            if not isinstance(s, dict):
                continue
            cnum = (s.get("container_number") or "").upper()
            if not cnum:
                continue
            carrier = s.get("carrier") or {}
            carrier_name = carrier.get("name", "") if isinstance(carrier, dict) else str(carrier)
            db[cnum] = {
                "container_number": cnum,
                "bl_number": "",
                "shipping_line": carrier_name,
                "carrier_scac": carrier.get("scac", "") if isinstance(carrier, dict) else "",
                "reference": s.get("reference", "") or "",
                "shipment_id": s.get("id", ""),
                "registered_at": s.get("created_at", ""),
                "last_refreshed": None,
            }
        save_tracking_db(db)
        logger.info(f"  Populated DB with {len(db)} shipments from API")

    for key, record in db.items():
        sid = str(record.get("shipment_id", ""))
        cnum = record.get("container_number", "").upper()
        shipment = shipment_map.get(sid) or shipment_map.get(cnum)

        if shipment:
            full_id = shipment.get("id")
            if full_id:
                try:
                    shipment = client.get_shipment(full_id)
                    record["shipment_id"] = full_id
                except Exception:
                    pass
            fields = extract_fields(shipment)
            record.update(fields)
            record["voyage_data"] = shipment
            record["last_refreshed"] = datetime.now(timezone.utc).isoformat()
            logger.info(f"  {key}: Status={fields['status']}, ETA={fields['eta']}, "
                        f"{fields['pol']} -> {fields['pod']}")
        else:
            logger.warning(f"  {key}: No matching shipment found")
            record["last_refreshed"] = datetime.now(timezone.utc).isoformat()
        save_tracking_db(db)
    return db


# ---------------------------------------------------------------------------
# Excel Export
# ---------------------------------------------------------------------------
def export_to_excel(db, output_path=OUTPUT_XLSX):
    if not HAS_OPENPYXL:
        logger.error("openpyxl not installed. Run: pip install openpyxl")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Container Tracking"

    hfont = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    status_colors = {
        "sailing": "D6EAF8", "en_route": "D6EAF8",
        "arrived": "D5F5E3", "discharged": "ABEBC6",
        "delivered": "82E0AA", "gate_out": "82E0AA",
        "booked": "FCF3CF", "new": "FCF3CF",
        "untracked": "F2F3F4",
    }

    headers = ["Container #", "Carrier", "Status", "Transit %",
               "ETD", "Port of Loading", "ETA", "Port of Discharge",
               "Vessel", "Last Refreshed"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = halign
        cell.border = border

    row = 2
    for key, rec in sorted(db.items()):
        transit_pct = rec.get("transit_pct", "")
        if transit_pct != "":
            transit_pct = f"{transit_pct}%"
        vals = [
            rec.get("container_number") or rec.get("bl_number") or key,
            rec.get("carrier", rec.get("shipping_line", "")),
            rec.get("status", "Pending"),
            transit_pct,
            rec.get("etd", ""),
            rec.get("pol", ""),
            rec.get("eta", ""),
            rec.get("pod", ""),
            rec.get("vessel", ""),
            rec.get("last_refreshed", "")[:16] if rec.get("last_refreshed") else "",
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="center")

        sc = ws.cell(row=row, column=3)
        sl = str(sc.value).lower().replace(" ", "_")
        for sk, color in status_colors.items():
            if sk in sl:
                sc.fill = PatternFill(start_color=color, fill_type="solid")
                break
        row += 1

    for col in range(1, len(headers) + 1):
        ml = max(len(str(ws.cell(row=r, column=col).value or ""))
                 for r in range(1, row))
        ws.column_dimensions[get_column_letter(col)].width = min(ml + 4, 35)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"
    wb.save(output_path)
    logger.info(f"Excel report saved to: {output_path}")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Container ETA Tracker (ShipsGo API v2)")
    parser.add_argument("--init", action="store_true", help="Create sample containers.json")
    parser.add_argument("--add", nargs=2, metavar=("CONTAINER", "LINE"))
    parser.add_argument("--ref", default="")
    parser.add_argument("--refresh", action="store_true")
    parser.add_argument("--export", action="store_true")
    parser.add_argument("--carriers", action="store_true")
    parser.add_argument("--all", action="store_true")
    args = parser.parse_args()

    if args.init:
        save_json(CONTAINERS_FILE, [
            {"container_number": "MSKU1234567", "shipping_line": "MAERSK LINE"},
        ])
        logger.info(f"Created sample {CONTAINERS_FILE}")
        return

    api_key = get_api_key()
    client = ShipsGoClient(api_key)

    if args.carriers:
        carriers = client.get_carriers()
        print("\nSupported Carriers:")
        print("-" * 50)
        for c in (carriers if isinstance(carriers, list) else [carriers]):
            if isinstance(c, dict):
                print(f"  {c.get('scac', '????')} - {c.get('name', 'Unknown')}")
            else:
                print(f"  {c}")
        return

    db = load_tracking_db()

    if args.add:
        cn, sl = args.add
        db = register_containers(client, [{"container_number": cn.upper(),
                                           "shipping_line": sl, "reference": args.ref}], db)
        db = refresh_all(client, db)
        export_to_excel(db)
        return

    if args.all or (not args.refresh and not args.export):
        containers = load_containers()
        if containers:
            db = register_containers(client, containers, db)
        db = refresh_all(client, db)
        export_to_excel(db)
        return

    if args.refresh:
        db = refresh_all(client, db)
        save_tracking_db(db)
    if args.export:
        export_to_excel(db)

if __name__ == "__main__":
    main()
