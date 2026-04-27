"""Read/write/template helpers for the user-linked Excel workbook."""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from container_tracker.core.constants import CONTAINER_COL_KEYWORDS
from container_tracker.core.util import now_est

TRACKING_COL_MAP = {
    "Carrier": "carrier",
    "Status": "status",
    "ETA": "eta",
    "Original ETA": "original_eta",
    "Delay": "delay_days",
    "Port of Loading": "pol",
    "Port of Discharge": "pod",
    "Vessel": "vessel",
    "Transit %": "transit_pct",
    "Last Refreshed": "last_refreshed",
}


def find_container_column(ws):
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(row=1, column=c).value or "").strip().lower()
        if h in CONTAINER_COL_KEYWORDS:
            return c
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(row=1, column=c).value or "").strip().lower()
        if "container" in h or "cntr" in h:
            return c
    return None


def find_or_create_tracking_columns(ws):
    existing = {}
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(row=1, column=c).value or "").strip()
        if h:
            existing[h.lower()] = c
    fm = {}
    nc = ws.max_column + 1
    hf = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    hfill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center")
    for hn, fk in TRACKING_COL_MAP.items():
        fc = existing.get(hn.lower())
        if fc:
            fm[fk] = fc
        else:
            c = ws.cell(row=1, column=nc, value=hn)
            c.font = hf
            c.fill = hfill
            c.alignment = ha
            fm[fk] = nc
            nc += 1
    return fm


def read_containers_from_excel(path):
    wb = load_workbook(str(path), data_only=True)
    ws = wb.active
    cc = find_container_column(ws)
    if cc is None:
        wb.close()
        return []
    out = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=cc).value
        if v:
            cn = str(v).strip().upper()
            if len(cn) >= 10:
                out.append(cn)
    wb.close()
    return out


def update_excel_with_tracking(path, data):
    wb = load_workbook(str(path))
    ws = wb.active
    cc = find_container_column(ws)
    if cc is None:
        wb.close()
        raise ValueError("No Container column found.")
    fm = find_or_create_tracking_columns(ws)
    sc = {"sailing": "D6EAF8", "en_route": "D6EAF8", "arrived": "D5F5E3",
          "discharged": "ABEBC6", "delivered": "82E0AA", "booked": "FCF3CF",
          "new": "FCF3CF", "untracked": "F2F3F4"}
    count = 0
    ts = now_est()
    for r in range(2, ws.max_row + 1):
        cv = ws.cell(row=r, column=cc).value
        if not cv:
            continue
        cn = str(cv).strip().upper()
        if cn in data:
            rec = data[cn]
            for fk, col in fm.items():
                val = rec.get(fk, "")
                if fk == "transit_pct" and val != "":
                    val = f"{val}%"
                if fk == "last_refreshed":
                    val = ts
                ws.cell(row=r, column=col, value=val)
            scol = fm.get("status")
            if scol:
                cell = ws.cell(row=r, column=scol)
                sl = str(cell.value or "").lower().replace(" ", "_")
                for sk, color in sc.items():
                    if sk in sl:
                        cell.fill = PatternFill(start_color=color, fill_type="solid")
                        break
            dcol = fm.get("delay_days")
            if dcol:
                dc = ws.cell(row=r, column=dcol)
                dv = str(dc.value or "")
                if dv.startswith("+"):
                    dc.fill = PatternFill(start_color="FADBD8", fill_type="solid")
                    dc.font = Font(color="C0392B")
                elif "early" in dv:
                    dc.fill = PatternFill(start_color="D5F5E3", fill_type="solid")
                    dc.font = Font(color="27AE60")
                elif "On time" in dv:
                    dc.font = Font(color="27AE60")
            count += 1

    existing_containers = set()
    for r in range(2, ws.max_row + 1):
        cv = ws.cell(row=r, column=cc).value
        if cv:
            existing_containers.add(str(cv).strip().upper())
    appended = 0
    for cn, rec in data.items():
        if cn not in existing_containers and rec.get("status"):
            nr = ws.max_row + 1
            ws.cell(row=nr, column=cc, value=cn)
            for fk, col in fm.items():
                val = rec.get(fk, "")
                if fk == "transit_pct" and val != "":
                    val = f"{val}%"
                if fk == "last_refreshed":
                    val = ts
                ws.cell(row=nr, column=col, value=val)
            appended += 1
            count += 1
    for fk, col in fm.items():
        ml = max((len(str(ws.cell(row=r, column=col).value or ""))
                  for r in range(1, ws.max_row + 1)), default=10)
        ws.column_dimensions[get_column_letter(col)].width = min(ml + 4, 30)
    wb.save(str(path))
    wb.close()
    return count


def append_container_row(path, cn, carrier="", status="NEW"):
    """Append a single new CN to the linked workbook. Idempotent — if cn
    is already in the container column, returns False without writing.

    Used by add_container so a freshly-tracked CN appears in the user's
    spreadsheet immediately, before the next refresh. Only the carrier
    and status columns are populated; tracked fields the bridge doesn't
    have yet (eta, route, etc.) stay blank and get filled on first
    refresh by update_excel_with_tracking.
    """
    cn_up = str(cn).strip().upper()
    if not cn_up:
        return False
    wb = load_workbook(str(path))
    try:
        ws = wb.active
        cc = find_container_column(ws)
        if cc is None:
            raise ValueError("No Container column found.")
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=cc).value
            if v and str(v).strip().upper() == cn_up:
                return False
        fm = find_or_create_tracking_columns(ws)
        nr = ws.max_row + 1
        ws.cell(row=nr, column=cc, value=cn_up)
        if carrier and "carrier" in fm:
            ws.cell(row=nr, column=fm["carrier"], value=carrier)
        if status and "status" in fm:
            ws.cell(row=nr, column=fm["status"], value=status)
        wb.save(str(path))
        return True
    finally:
        wb.close()


def remove_container_row(path, cn):
    """Delete the workbook row(s) whose Container # cell matches cn.
    No-op if cn isn't found. Returns the count of rows removed.

    Used by archive_container to keep the spreadsheet in sync when the
    user archives a container in the UI.
    """
    cn_up = str(cn).strip().upper()
    if not cn_up:
        return 0
    wb = load_workbook(str(path))
    try:
        ws = wb.active
        cc = find_container_column(ws)
        if cc is None:
            return 0
        # Walk top-down to collect, delete bottom-up so indices don't shift.
        to_delete = []
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=cc).value
            if v and str(v).strip().upper() == cn_up:
                to_delete.append(r)
        for r in reversed(to_delete):
            ws.delete_rows(r)
        if to_delete:
            wb.save(str(path))
        return len(to_delete)
    finally:
        wb.close()


def create_template_excel(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Container Tracking"
    headers = ["Container #", "PO / Reference", "Notes", "Carrier", "Status", "ETA",
               "Original ETA", "Delay", "Port of Loading", "Port of Discharge",
               "Vessel", "Transit %", "Last Refreshed"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    for ri, (cn, ref, n) in enumerate(
        [("FAKE0000001", "PO-SAMPLE-001", "Replace with your data"),
         ("FAKE0000002", "PO-SAMPLE-002", "Replace with your data")], 2):
        ws.cell(row=ri, column=1, value=cn)
        ws.cell(row=ri, column=2, value=ref)
        ws.cell(row=ri, column=3, value=n)
    lc = get_column_letter(len(headers))
    tbl = Table(displayName="ContainerTracking", ref=f"A1:{lc}3")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)
    for i, w in enumerate([18, 18, 25, 16, 14, 14, 14, 14, 20, 20, 20, 12, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(str(path))
    wb.close()
    return str(path)
