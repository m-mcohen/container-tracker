"""Shared pytest fixtures.

* ``isolated_data_dir`` redirects DATA_DIR / CONFIG_FILE / TRACKING_DB_FILE /
  LOG_FILE inside ``tmp_path`` so config and migration tests don't touch the
  real %APPDATA%\\ContainerTracker folder.

* ``mock_keyring`` replaces ``container_tracker.core.credentials.keyring`` with
  an in-memory dict-backed mock. Per-test isolation; no real Credential Manager
  reads/writes.

* ``sample_workbook`` builds a Container-tracking xlsx in ``tmp_path`` on demand
  so tests stay reviewable in diff (no committed binary).
"""

from __future__ import annotations

from pathlib import Path
from typing import Callable

import pytest


@pytest.fixture
def isolated_data_dir(tmp_path, monkeypatch):
    """Point all DATA_DIR-derived paths at tmp_path. Must be requested BEFORE
    any test code calls config.boot() / load_config() / save_config()."""
    from container_tracker.core import config as ct_config

    monkeypatch.setattr(ct_config, "DATA_DIR", tmp_path)
    monkeypatch.setattr(ct_config, "CONFIG_FILE", tmp_path / "config.json")
    monkeypatch.setattr(ct_config, "TRACKING_DB_FILE", tmp_path / "tracking_data.json")
    monkeypatch.setattr(ct_config, "LOG_FILE", tmp_path / "tracker.log")
    return tmp_path


class _FakeKeyring:
    """Drop-in replacement for the keyring module. Stores secrets in a dict
    keyed by (service, user). Tracks call counts so idempotency tests can
    assert no-churn."""

    def __init__(self):
        self.store: dict[tuple[str, str], str] = {}
        self.set_calls: list[tuple[str, str, str]] = []
        self.delete_calls: list[tuple[str, str]] = []
        self.get_calls: list[tuple[str, str]] = []

    def get_password(self, service, user):
        self.get_calls.append((service, user))
        return self.store.get((service, user))

    def set_password(self, service, user, password):
        self.set_calls.append((service, user, password))
        self.store[(service, user)] = password

    def delete_password(self, service, user):
        self.delete_calls.append((service, user))
        self.store.pop((service, user), None)

    def reset_call_counts(self):
        self.set_calls.clear()
        self.delete_calls.clear()
        self.get_calls.clear()


@pytest.fixture
def mock_keyring(monkeypatch):
    """Replace credentials.keyring with an in-memory fake."""
    from container_tracker.core import credentials

    fake = _FakeKeyring()
    monkeypatch.setattr(credentials, "keyring", fake)
    return fake


@pytest.fixture
def sample_tracking_db(isolated_data_dir):
    """Factory: write a tracking_data.json into the isolated data dir and
    return the loaded dict. Used by bridge tests that need realistic
    on-disk records.

    Default record is one fully-refreshed entry sourced from the
    sanitized shipsgo_response.json fixture: extract_fields() runs on
    the payload and the result is merged into the record (mirrors the
    legacy GUI's ``rec.update(extract_fields(sh))`` write path)."""
    import json
    from container_tracker.core import config as ct_config
    from container_tracker.core.status import extract_fields

    fixtures = Path(__file__).parent / "fixtures"

    def make(records: dict | None = None) -> dict:
        if records is None:
            shipment = json.loads((fixtures / "shipsgo_response.json").read_text())
            inner = shipment["shipment"]
            extracted = extract_fields(shipment)
            cn = inner["container_number"].upper()
            records = {
                cn: {
                    "container_number": cn,
                    "shipment_id": inner["id"],
                    "last_refreshed": "2026-04-25 10:00 AM EST",
                    **extracted,
                }
            }
        ct_config.TRACKING_DB_FILE.write_text(json.dumps(records, default=str))
        return records

    return make


@pytest.fixture
def sample_workbook(tmp_path) -> Callable[..., Path]:
    """Factory: returns a callable ``make(headers=..., rows=...)`` that writes a
    1-sheet xlsx into tmp_path and returns its Path. Default headers expose
    only the canonical 'Container #' column; tests pass alternative headers
    (e.g. 'cntr no') by overriding."""
    from openpyxl import Workbook

    def make(
        headers: tuple[str, ...] = ("Container #",),
        rows: tuple[str, ...] = ("MSCU1111111", "MAEU2222222", "EGLV3333333"),
        filename: str = "containers.xlsx",
    ) -> Path:
        wb = Workbook()
        ws = wb.active
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        for ri, cn in enumerate(rows, 2):
            ws.cell(row=ri, column=1, value=cn)
        path = tmp_path / filename
        wb.save(str(path))
        return path

    return make
