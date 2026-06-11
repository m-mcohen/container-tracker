"""Unit tests for container_tracker.core.excel."""

import pytest
from openpyxl import load_workbook

from container_tracker.core.excel import (
    find_container_column,
    read_containers_from_excel,
    update_excel_with_tracking,
)


class TestReadContainersFromExcel:
    def test_returns_uppercased_normalized_list(self, sample_workbook):
        path = sample_workbook(rows=("mscu1111111", "  MAEU2222222  ", "EGLV3333333"))
        result = read_containers_from_excel(path)
        assert result == ["MSCU1111111", "MAEU2222222", "EGLV3333333"]

    def test_skips_short_values(self, sample_workbook):
        # Anything under 10 chars is treated as noise (e.g. notes / IDs).
        path = sample_workbook(rows=("MSCU1111111", "X", "AB", "EGLV3333333"))
        result = read_containers_from_excel(path)
        assert result == ["MSCU1111111", "EGLV3333333"]

    def test_alternative_header_cntr_no(self, sample_workbook):
        path = sample_workbook(headers=("cntr no",), rows=("MSCU9999999",))
        assert read_containers_from_excel(path) == ["MSCU9999999"]

    def test_returns_empty_when_no_container_column(self, sample_workbook):
        path = sample_workbook(headers=("UnrelatedHeader",), rows=("MSCU1111111",))
        assert read_containers_from_excel(path) == []


class TestUpdateExcelWithTracking:
    def test_updates_existing_row_and_creates_tracking_columns(self, sample_workbook):
        path = sample_workbook(rows=("MSCU1111111",))
        data = {"MSCU1111111": {
            "carrier": "MSC", "status": "SAILING", "eta": "2026-07-01",
            "original_eta": "2026-06-28", "delay_days": "+3 days",
            "pol": "SHANGHAI", "pod": "LOS ANGELES", "vessel": "MSC OSCAR",
            "transit_pct": 40,
        }}

        count = update_excel_with_tracking(path, data)

        assert count == 1
        wb = load_workbook(str(path))
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value
                   for c in range(1, ws.max_column + 1)]
        assert "Status" in headers and "ETA" in headers and "Vessel" in headers
        assert ws.cell(row=2, column=headers.index("Status") + 1).value == "SAILING"
        assert ws.cell(row=2, column=headers.index("Transit %") + 1).value == "40%"
        wb.close()

    def test_raises_when_no_container_column(self, sample_workbook):
        # The bridge relies on this raising (→ excel_write_failed banner);
        # a silent no-op here would hide a renamed-header workbook.
        path = sample_workbook(headers=("UnrelatedHeader",), rows=("MSCU1111111",))
        with pytest.raises(ValueError):
            update_excel_with_tracking(path, {"MSCU1111111": {"status": "SAILING"}})

    def test_appends_row_for_db_only_container_with_status(self, sample_workbook):
        path = sample_workbook(rows=("MSCU1111111",))
        data = {
            "MSCU1111111": {"status": "SAILING"},
            "EGLV9999999": {"status": "ARRIVED"},
        }

        count = update_excel_with_tracking(path, data)

        assert count == 2
        assert "EGLV9999999" in read_containers_from_excel(path)

    def test_statusless_stub_not_appended(self, sample_workbook):
        # Pre-refresh stubs (no API data yet) must not be written into the
        # user's workbook — they'd appear as blank rows.
        path = sample_workbook(rows=("MSCU1111111",))
        data = {"NEWU1234567": {"container_number": "NEWU1234567"}}

        assert update_excel_with_tracking(path, data) == 0
        assert "NEWU1234567" not in read_containers_from_excel(path)


class TestFindContainerColumn:
    def test_canonical_header_index_1(self, sample_workbook):
        path = sample_workbook(headers=("Container #",))
        wb = load_workbook(str(path))
        assert find_container_column(wb.active) == 1
        wb.close()

    def test_substring_match_fallback(self, sample_workbook):
        # "Master container" doesn't match the keyword list literally, but the
        # substring 'container' triggers the second-pass fallback.
        path = sample_workbook(headers=("Master container",), rows=("MSCU1111111",))
        wb = load_workbook(str(path))
        assert find_container_column(wb.active) == 1
        wb.close()
